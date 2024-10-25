import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment

# URL of the website to scrape
base_url = 'https://innovatoor.com/'

def fetch_page(url):
    """Function to fetch the HTML content of a webpage."""
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.content
        else:
            print(f'Failed to fetch page. Status code: {response.status_code}')
    except requests.exceptions.RequestException as e:
        print(f'Error fetching page: {str(e)}')
    return None

def parse_html(content):
    """Function to parse HTML content using BeautifulSoup."""
    osint_content = []
    if content:
        soup = BeautifulSoup(content, 'html.parser')
        # Example: Extracting paragraphs containing 'OSINT'
        paragraphs = soup.find_all('p', text=lambda text: text and 'OSINT' in text)
        for p in paragraphs:
            osint_content.append(p.text.strip())

        # Example: Extracting headings containing 'OSINT'
        headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'], text=lambda text: text and 'OSINT' in text)
        for heading in headings:
            osint_content.append(heading.text.strip())

        # Add more specific parsing logic as per the website's structure

    return osint_content

def save_to_excel(content):
    """Function to save scraped content to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'OSINT Content'

    # Set column widths and alignment
    ws.column_dimensions['A'].width = 100
    ws['A1'].alignment = Alignment(wrap_text=True)

    # Write headers
    ws['A1'] = 'OSINT Content'

    # Write content
    for idx, item in enumerate(content, start=2):
        ws.cell(row=idx, column=1).value = item
        ws.cell(row=idx, column=1).alignment = Alignment(wrap_text=True)

    # Save workbook
    wb.save('osint_content.xlsx')
    print('OSINT content saved to osint_content.xlsx')

def scrape_osint_content():
    """Function to initiate scraping of pages containing 'OSINT'."""
    url = base_url
    html_content = fetch_page(url)
    if html_content:
        osint_content = parse_html(html_content)
        if osint_content:
            save_to_excel(osint_content)
        else:
            print('No OSINT content found on the page.')
    else:
        print('Failed to retrieve content.')

if __name__ == '__main__':
    scrape_osint_content()
